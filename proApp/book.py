# -*-  coding:utf-8 -*-
from django.shortcuts import render
from django.http import HttpResponse
from django.core import serializers
from django.core.files.base import ContentFile
import json,time
from proApp import models, base, common
import logging
logger = logging.getLogger('django')

# from proApp import models          # 导出Excel import
from django.http import HttpResponse
import xlwt
from io import BytesIO
import os                         # 导出Excel import

from openpyxl import Workbook,load_workbook  # 导入 excel
from openpyxl.utils import get_column_letter
from openpyxl.compat import range    # 导入 excel
import xlrd     #excel读工具

import random,imghdr,shutil          # 头像上传

@base.checkLogin
def index(request):
    return render(request,"book.html")

def getBook(request):
    if request.method == "GET":
        page = int(request.GET.get('page',0))-1
        rows = int(request.GET.get('rows',-1))
        list = models.Book.objects.all().order_by("-publish_date")
        allList = bookInfo(list)
        total = len(allList)
        if page < 0 or rows == 0:
            json_data_list = allList
        else:
            p = common.Datagrid()
            json_data_list = p.page(page, rows, total, allList)
        return HttpResponse(json.dumps(json_data_list), content_type='application/json; charset=utf-8')

    if request.method == "POST":
        page = int(request.POST.get('page', '')) - 1
        rows = int(request.POST.get('rows', ''))
        name = request.POST.get("name", '')
        price = request.POST.get("price")
        publisher = request.POST.get("publisher")
        author = request.POST.get("author")
        dateFrom = request.POST.get("dateFrom")
        dateTo = request.POST.get("dateTo")

        # 定一个字典用于保存前端发送过来的查询条件
        search_dict = dict()

        if dateFrom:
            search_sql = models.Book.objects.filter(publish_date__gte=dateFrom)
        else:
            search_sql = models.Book.objects

        if dateTo:
            search_sql = search_sql.filter(publish_date__lte=dateTo)

        if publisher and int(publisher)>-1:
            search_dict['publisher_id'] = publisher

        if author and int(author)>-1:
            search_dict['author'] = author

        # 序列化
        if name!="全部":
            list = search_sql.filter(name__contains=name).filter(**search_dict).order_by("-publish_date")
        else:
            list = search_sql.filter(**search_dict).order_by("-publish_date")

        allList = bookInfo(list)
        total = len(allList)
        p = common.Datagrid()
        json_data_list = p.page(page, rows, total, allList)
        return HttpResponse(json.dumps(json_data_list), content_type='application/json; charset=utf-8')

#  因获取后台数据前端不能直接使用，进行重组， 接收参数list: 数据库所有符合条件的数据
def bookInfo(list):
    allList = []
    for li in list:
        book_list = json.loads(serializers.serialize("json", li.author.all(), ensure_ascii=False))
        author_name = ''
        author_id = []
        author = []
        for i,a in enumerate(book_list):
            author_id.append(a['pk'])
            author_name += a['fields']['name'] + '、'
            author.append({"id":a['pk'],"name":a['fields']['name']})

        publisher = json.loads(serializers.serialize("json", models.Publisher.objects.filter(id=li.publisher_id), ensure_ascii=False))
        allList.append({
            "id": li.id,
            # 'bookCover': li.bookCover,
            "name": li.name,
            "price": li.price,
            "saleNum": li.saleNum,
            "publisher": li.publisher_id,
            "publisher_str": publisher[0]['fields']['name'],
            "author_id": author_id,
            "author_name": author_name,
            "author": author_name,
            "publish_date": json.loads(json.dumps(li.publish_date, cls=common.DateEncoder)),
        })
    return allList

def addBook(request):
    name = request.POST.get("name")
    price = request.POST.get("price")
    saleNum = request.POST.get("saleNum")
    publisher = models.Publisher.objects.get(pk=request.POST.get("publisher"))
    authors_list = request.POST.getlist("author[]")

    book = models.Book.objects.all()
    if len(book) > 0:
        maxId = models.Book.objects.latest('id').id
    else:
        maxId = 0
    maxId += 1

    dic = {
        'id': maxId,
        'name': name,
        'price': price,
        'saleNum': saleNum,
        "publisher": publisher,
        # "publish_date": json.loads(json.dumps(time.time(), cls=DateEncoder)),
    }
    b1 = models.Book(**dic)
    b1.save()                                     # 普通插入的数据和外键插入的数据需要先save()
    b1 = models.Book.objects.get(id=maxId)       # 查出书名对象,也就是获取要插入的多对多数据项
    if len(authors_list) == 1:
        b1.author.add(authors_list[0])  # 多对多使用add方法进行插入
        b1.save()
    else:
        for person in authors_list:     # 循环插入用户选中的多个作者
            b1.author.add(person)        # 多对多使用add方法进行插入
        b1.save()

    ret = {
        'success': True,
        'retCode': 200,
        'retMsg': "添加成功！"
    }
    return HttpResponse(json.dumps(ret), content_type='application/json')


def addBook2(request):
    name = request.POST.get("name")
    price = request.POST.get("price")
    saleNum = request.POST.get("saleNum")
    publisher = models.Publisher.objects.get(pk=request.POST.get("publisher"))
    authors_list = request.POST.getlist("author[]")
    dic = {
        'name': name,
        'price': price,
        'saleNum': saleNum,
        "publisher": publisher,
        # "publish_date": json.loads(json.dumps(time.time(), cls=DateEncoder)),
    }
    b1 = models.Book(**dic)
    b1.save()                                     # 普通插入的数据和外键插入的数据需要先save()
    b1 = models.Book.objects.get(name=name)       # 查出书名对象,也就是获取要插入的多对多数据项
    if len(authors_list) == 1:
        b1.author.add(authors_list[0])  # 多对多使用add方法进行插入
        b1.save()
    else:
        for person in authors_list:     # 循环插入用户选中的多个作者
            b1.author.add(person)        # 多对多使用add方法进行插入
        b1.save()
    ret = {
        'success': True,
        'retCode': 200,
        'retMsg': "添加成功！"
    }
    return HttpResponse(json.dumps(ret), content_type='application/json')

def delBook(request):
    id = request.POST.get("id")
    models.Book.objects.filter(id=id).delete()
    ret = {
        'success': True,
        'retCode': 0,
        'retMsg': "删除成功！"
    }
    return HttpResponse(json.dumps(ret), content_type='application/json')

def modifyBook(request):
    try:
        id = request.POST.get("modifyId")
        name = request.POST.get("name")
        price = request.POST.get("price")
        saleNum = request.POST.get("saleNum")
        publisher = request.POST.get("publisher")
        author_list = request.POST.getlist("author[]")
        # bookCover = request.FILES.get("bookCover")
        # bookCover = request.user.bookCover
        # print (bookCover)
        # fileMore = request.FILES.get("moreFile")
        models.Book.objects.filter(id=id).update(name=name, price=float(price), saleNum=saleNum, publisher=publisher)
        book_obj = models.Book.objects.get(id=id)
        author_obj = models.Author.objects.filter(id__in=author_list)
        book_obj.author.set(author_obj)
        book_obj.save()

        ret = {
            'success': True,
            'retCode': 200,
            'retMsg': "修改成功！"
        }
    except Exception as e:
        logger.info('-------------------------')
        logger.error(str(e))
        logger.warn('warn')
        logger.debug('debug')
        ret = {
            'success': False,
            'retCode': 200,
            'retMsg': str(e)
        }

    return HttpResponse(json.dumps(ret), content_type='application/json')

#  表格导出功能
def exportBookExcel(request):
    list_obj = models.Book.objects.all().order_by("-publish_date")
    if list_obj:
        # 创建工作薄
        ws = xlwt.Workbook(encoding='utf-8')
        w = ws.add_sheet(u"sheet1",u"sheet2")
        w.write(0, 0, "id")
        w.write(0, 1, u"名称")
        w.write(0, 2, u"价格")
        w.write(0, 3, u"销量")
        w.write(0, 4, u"出版社id")
        w.write(0, 5, u"出版社")
        w.write(0, 6, u"作者")
        w.write(0, 7, u"发布时间")
        # 写入数据
        excel_row = 1
        for obj in list_obj:
            data_id = obj.id
            data_name = obj.name
            data_price = obj.price
            data_saleNum = obj.saleNum
            data_publisher = str(obj.publisher)
            data_publisher_id = obj.publisher_id
            data_author = obj.author_list()
            data_time = obj.publish_date.strftime("%Y-%m-%d %H:%M:%S")[:20]
            w.write(excel_row, 0, data_id)
            w.write(excel_row, 1, data_name)
            w.write(excel_row, 2, data_price)
            w.write(excel_row, 3, data_saleNum)
            w.write(excel_row, 4, data_publisher_id)
            w.write(excel_row, 5, data_publisher)
            w.write(excel_row, 6, data_author)
            w.write(excel_row, 7, data_time)
            excel_row += 1
        # 检测文件是够存在
        # 方框中代码是保存本地文件使用，如不需要请删除该代码
        ###########################
        exist_file = os.path.exists("book.xls")
        if exist_file:
            os.remove(r"book.xls")
        ws.save("book.xls")
        ############################
        sio = BytesIO()
        ws.save(sio)
        sio.seek(0)
        response = HttpResponse(sio.getvalue(), content_type='application/vnd.ms-excel')
        response['Content-Disposition'] = 'attachment; filename=book.xls'
        response.write(sio.getvalue())
        return response

# 表格导入功能
def importBook(request):
    """
        加载导入的excel数据
        :param path: 文件地址
        :param sheet_name: 获取指定sheet，默认加载全部
        :param ignore_header: 是否忽略表头数据，即第一行，默认否
        :return: 返回数组集
    """
    if request.method == 'GET':
        return render(request, 'index.html')
    elif request.method == 'POST':
        f = request.FILES.get('bookExcel', None)
        if not f:
            logger.error('no files for upload!')
            ret = {
                'success': True,
                'retCode': 200,
                'retMsg': "no files for upload!"
            }
            return HttpResponse(json.dumps(ret), content_type='application/json')

        excel_type = f.name.split('.')[1]
        if excel_type in ['xlsx','xls']:
            # 开始解析上传的excel表格
            wb = xlrd.open_workbook(filename=None,file_contents=f.read())
            table = wb.sheets()[0]
            rows = table.nrows  # 总行数
            try:
                # with transaction.atomic():  # 控制数据库事务交易
                #     for i in range(1,rows):
                #         rowVlaues = table.row_values(i)
                #         major = models.Book.objects.filter(majorid=rowVlaues[1]).first()
                #         models.Book.objects.create(gradeid=rowVlaues[0],major=major,gradename=rowVlaues[2],memo=rowVlaues[3])
                for i in range(1, rows):
                    rowVlaues = table.row_values(i)
                    publisher = models.Publisher.objects.get(pk=str(int(rowVlaues[3])))   # 根据id插入
                    # publisher = models.Publisher.objects.get(name=rowVlaues[4])             # 根据name插入  Publisher表不允许有重复名

                    # 手动获取表的主键，数据插入时进行最大赋值
                    book = models.Book.objects.all()
                    if len(book) > 0:
                        maxId = models.Book.objects.latest('id').id
                    else:
                        maxId = 0
                    maxId += 1

                    dic = {
                        'id': maxId,
                        'name': rowVlaues[0],
                        'price': str(round(rowVlaues[1],2)),
                        'saleNum': str(int(rowVlaues[2])),
                        "publisher": publisher
                    }
                    b1 = models.Book(**dic)
                    b1.save()                                        # 普通插入的数据和外键插入的数据需要先save()
                    b1 = models.Book.objects.get(id=maxId)           # 查出书名对象,也就是获取要插入的多对多数据项
                    authors_list = []
                    for authorId in rowVlaues[5].split(","):
                        authors_list.append(models.Author.objects.get(id=authorId))

                    if len(authors_list) == 1:
                        b1.author.add(authors_list[0])  # 多对多使用add方法进行插入
                        b1.save()
                    else:
                        for person in authors_list:    # 循环插入用户选中的多个作者
                            b1.author.add(person)        # 多对多使用add方法进行插入
                        b1.save()
                ret = {
                    'success': True,
                    'retCode': 200,
                    'retMsg': "导入成功！"
                }
                return HttpResponse(json.dumps(ret), content_type='application/json')
            except:
                logger.error('解析excel文件或者数据插入错误')
                print (rows)
                ret = {
                    'success': False,
                    'retCode': 200,
                    'retMsg': "解析excel文件或者数据插入错误！"
                }
                return HttpResponse(json.dumps(ret), content_type='application/json')
        else:
            logger.error('上传文件类型错误！')
            ret = {
                'success': True,
                'retCode': 200,
                'retMsg': "导入失败！"
            }
            return HttpResponse(json.dumps(ret), content_type='application/json')

#  图书封面上传
def uploadBookCover(request):
    root_path = 'media/img'
    if request.method == 'POST':
        bookCover = request.FILES.get('bookCover')
        userId = request.user.id
        path = '%s/%s_bookCover'%(root_path,userId)                # 设定图片路径
        bookRet =  handleUpdateBookCover(bookCover,path,userId)      # 获得bookRet，handleUpdateBookCover处理
        if bookRet:
            request.user.bookCover = '%s/%s' % (path,bookRet[1])
            request.user.save()
            ret = {
                'success': True,
                'retCode': 200,
                'retMsg': "图书封面上传成功！"
            }
            return HttpResponse(json.dumps(ret), content_type='application/json')
        ret = {
            'success': False,
            'retCode': 200,
            'retMsg': "图书封面上传失败！"
        }
        return HttpResponse(json.dumps(ret), content_type='application/json')

def handleUpdateBookCover(bookCover,path,userId):
    """  return bookCover full path if success """
    if bookCover.size>=5*1024*1024:
        return false
    bookCover_type_name_path = handleUpdateBookCoverFinally(bookCover)   # 对头像进行处理,返回一个List
    if not bookCover_type_name_path:
        return false
    try:
        if not os.path.exists(path):    # 假设路径不存在，创建目录
            os.mkdir(path)
        temp_path_fixed_name = 'media/img/bookCover/%s.%s'% (bookCover_type_name_path[1],bookCover_type_name_path[0])  #对文件重命名
        os.rename(bookCover_type_name_path[2],temp_path_fixed_name)
        listdir = os.listdir(path)
        if listdir:                                 # 移除原先头像
            os.remove(path+'/'+listdir[0])
        shutil.move(temp_path_fixed_name,path)     # 将新文件拷贝过去
        finally_book_name = bookCover_type_name_path[1]+"."+bookCover_type_name_path[0]
        return "media/img/%s_bookCover/%s.%s" % (userId,bookCover_type_name_path[1],bookCover_type_name_path[0]),finally_book_name
    except Exception as e:
        logger.error("上传图书封面报错：", str(e))
        ret = {
            'success': False,
            'retCode': 200,
            'retMsg': "图书封面上传失败！错误原因：" + str(e)
        }
        return HttpResponse(json.dumps(ret), content_type='application/json')


def handleUpdateBookCoverFinally(bookCover):
    temp_name = str(time.time() + random.randint(1,500))
    temp_path = 'media/img/bookCover/%s' % temp_name
    temp_file = open(temp_path,'wb')
    for chunk in bookCover.chunks():
        temp_file.write(chunk)          # 将头像上传到这个路径
    temp_file.close()
    img_type = imghdr.what(temp_path)   # 获取上传图片类型例如jpg,png等
    if not img_type:                   # 假如不是图片类型文件，移除文件并报错
        os.remove(temp_path)
        return false
    return img_type,temp_name,temp_path

#  单文件上传
def uploadFile(request):
    if request.method == 'POST':
        try:
            BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            file_obj = request.FILES.get('file', None)
            if not file_obj:
                ret = {
                    'success': False,
                    'retCode': 500,
                    'retMsg': "no files for upload!"
                }
                return HttpResponse(json.dumps(ret), content_type='application/json')

            dateTime = time.strftime('%Y%m%d',time.localtime(time.time()))
            if not os.path.exists(BASE_DIR+'/media/files/'+dateTime):
                os.mkdir(BASE_DIR + '/media/files/' + dateTime)
            f = open(os.path.join(BASE_DIR, 'media', 'files', dateTime, file_obj.name), 'wb')

            for chunk in file_obj.chunks():
                f.write(chunk)
            f.close()
            ret = {
                'success': True,
                'retCode': 200,
                'retMsg': "附件上传成功！"
            }
            return HttpResponse(json.dumps(ret), content_type='application/json')
        except Exception as e:
            logger.error('附件上传失败！错误原因：', str(e))
            ret = {
                'success': False,
                'retCode': 200,
                'retMsg': "附件上传失败！错误原因："+str(e)
            }
            return HttpResponse(json.dumps(ret), content_type='application/json')

#  多文件上传
def uploadMoreFile(request):
    if request.method == 'POST':
        try:
            BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
            file_obj = request.FILES.getlist('moreFile')
            if not file_obj:
                ret = {
                    'success': False,
                    'retCode': 500,
                    'retMsg': "no files for upload!"
                }
                return HttpResponse(json.dumps(ret), content_type='application/json')

            dateTime = time.strftime('%Y%m%d',time.localtime(time.time()))
            if not os.path.exists(BASE_DIR+'/media/files/'+dateTime):
                os.mkdir(BASE_DIR + '/media/files/' + dateTime)
            for file in file_obj:
                f = open(os.path.join(BASE_DIR, 'media', 'files', dateTime, file.name), 'wb')

                for chunk in file.chunks():
                    f.write(chunk)
                f.close()
            ret = {
                'success': True,
                'retCode': 200,
                'retMsg': "附件上传成功！"
            }
            return HttpResponse(json.dumps(ret), content_type='application/json')
        except Exception as e:
            logger.error('附件上传失败！错误原因：', str(e))
            ret = {
                'success': False,
                'retCode': 200,
                'retMsg': "附件上传失败！错误原因："+str(e)
            }
            return HttpResponse(json.dumps(ret), content_type='application/json')